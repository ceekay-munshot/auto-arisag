from __future__ import annotations

import json
import os
import re
import subprocess
from dataclasses import dataclass
from datetime import UTC, date, datetime
from pathlib import Path
import tempfile
from typing import Any
from urllib.parse import urljoin
from urllib.parse import urlencode

import requests
from openpyxl import load_workbook

VAHAN_ANALYTICS_URL = "https://analytics.parivahan.gov.in/analytics/publicdashboard/vahan?lang=en"
VAHAN_REPORT_URL = "https://vahan.parivahan.gov.in/vahan4dashboard/vahan/view/reportview.xhtml"

DEFAULT_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0 Safari/537.36"
    ),
    "Accept-Language": "en-IN,en;q=0.9",
}

MONTH_NAME_TO_NUMBER = {
    "JAN": 1,
    "FEB": 2,
    "MAR": 3,
    "APR": 4,
    "MAY": 5,
    "JUN": 6,
    "JUL": 7,
    "AUG": 8,
    "SEP": 9,
    "OCT": 10,
    "NOV": 11,
    "DEC": 12,
}

SEGMENT_FILTERS = {
    "PV": {
        "label": "Passenger Vehicles",
        "group": "main",
        "class_codes": ["5", "7", "15", "23", "31"],
    },
    "2W": {
        "label": "Two-Wheelers",
        "group": "main",
        "class_codes": ["1", "2", "3", "4", "5", "51", "52", "53"],
    },
    "3W": {
        "label": "Three-Wheelers",
        "group": "main",
        "class_codes": ["6", "54", "55", "57", "58"],
    },
    "CV": {
        "label": "Commercial Vehicles",
        "group": "main",
        "class_codes": [
            "8",
            "9",
            "10",
            "11",
            "12",
            "14",
            "16",
            "17",
            "18",
            "19",
            "20",
            "21",
            "24",
            "27",
            "28",
            "30",
            "32",
            "56",
            "59",
            "62",
            "64",
            "65",
            "66",
            "67",
            "68",
            "69",
            "70",
            "71",
            "73",
            "75",
            "76",
            "77",
            "78",
            "79",
            "80",
            "81",
            "82",
            "83",
            "84",
            "85",
            "86",
            "89",
            "91",
            "93",
            "94",
        ],
    },
    "TRACTOR": {
        "label": "Tractors",
        "group": "main",
        "class_codes": ["13", "63", "90"],
    },
    "CE": {
        "label": "Construction Equipment",
        "group": "main",
        "class_codes": ["22", "25", "26", "29", "87", "88", "92"],
    },
    "E2W": {
        "label": "EV 2W",
        "group": "ev",
        "class_codes": ["1", "2", "3", "4", "5", "51", "52", "53"],
        "fuel_codes": ["4"],
    },
    "E3W": {
        "label": "EV 3W",
        "group": "ev",
        "class_codes": ["6", "54", "55", "57", "58"],
        "fuel_codes": ["4"],
    },
    "EPV": {
        "label": "EV PV",
        "group": "ev",
        "class_codes": ["5", "7", "15", "23", "31"],
        "fuel_codes": ["4"],
    },
    "ECV": {
        "label": "EV CV",
        "group": "ev",
        "class_codes": [
            "8",
            "9",
            "10",
            "11",
            "12",
            "14",
            "16",
            "17",
            "18",
            "19",
            "20",
            "21",
            "24",
            "27",
            "28",
            "30",
            "32",
            "56",
            "59",
            "62",
            "64",
            "65",
            "66",
            "67",
            "68",
            "69",
            "70",
            "71",
            "73",
            "75",
            "76",
            "77",
            "78",
            "79",
            "80",
            "81",
            "82",
            "83",
            "84",
            "85",
            "86",
            "89",
            "91",
            "93",
            "94",
        ],
        "fuel_codes": ["4"],
    },
}


@dataclass
class VahanCacheResult:
    source: str
    status: str
    message: str
    updated: bool
    latest_value: str | None = None

    def to_dict(self) -> dict[str, str | bool | None]:
        return {
            "source": self.source,
            "status": self.status,
            "message": self.message,
            "updated": self.updated,
            "latest_value": self.latest_value,
            "checked_at": datetime.now(UTC).isoformat(),
        }


class VahanOemClient:
    def __init__(self) -> None:
        self.session = requests.Session()
        self.headers = dict(DEFAULT_HEADERS)
        self.ajax_headers = {
            **self.headers,
            "Faces-Request": "partial/ajax",
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
        }
        self.view_state = ""
        self.action_url = ""
        self.transport = "requests"
        self.cookie_jar_path: Path | None = None
        self._boot()

    def _boot(self) -> None:
        try:
            html = self._get_text(VAHAN_REPORT_URL, headers=self.headers, timeout=90)
        except Exception:
            self.transport = "curl"
            self.cookie_jar_path = Path(tempfile.NamedTemporaryFile(prefix="vahan_cookies_", suffix=".txt", delete=False).name)
            html = self._get_text(VAHAN_REPORT_URL, headers=self.headers, timeout=90)
        self.view_state = self._extract_view_state_from_html(html)
        self.action_url = self._extract_form_action(html)
        self._post_partial(
            source="yaxisVar",
            execute="yaxisVar",
            render="xaxisVar",
            pairs=[
                ("yaxisVar", "yaxisVar"),
                ("yaxisVar_input", "Maker"),
                ("xaxisVar_input", "VCG"),
                ("selectedYearType_input", "C"),
                ("selectedYear_input", str(date.today().year)),
            ],
        )
        self._post_partial(
            source="xaxisVar",
            execute="xaxisVar",
            render="multipleYear",
            pairs=[
                ("xaxisVar", "xaxisVar"),
                ("yaxisVar_input", "Maker"),
                ("xaxisVar_input", "Month Wise"),
                ("selectedYearType_input", "C"),
                ("selectedYear_input", str(date.today().year)),
            ],
        )

    def fetch_segment_year(self, segment_id: str, year: int) -> dict[str, Any]:
        segment = SEGMENT_FILTERS[segment_id]
        pairs = [
            ("j_idt32_input", "A"),
            ("j_idt41_input", "-1"),
            ("selectedRto_input", "-1"),
            ("yaxisVar_input", "Maker"),
            ("xaxisVar_input", "Month Wise"),
            ("selectedYearType_input", "C"),
            ("selectedYear_input", str(year)),
        ]
        for code in segment["class_codes"]:
            pairs.append(("VhClass", code))
        for fuel_code in segment.get("fuel_codes", []):
            pairs.append(("fuel", fuel_code))

        self._post_partial(
            source="j_idt73",
            execute="j_idt73 masterLayout_formlogin",
            render="VhCatg norms fuel VhClass combTablePnl groupingTable msg vhCatgPnl",
            pairs=[("j_idt73", "j_idt73"), *pairs],
            timeout=60,
        )

        workbook_content = self._post_bytes(
            self.action_url,
            headers=self.headers,
            pairs=[
                ("masterLayout_formlogin", "masterLayout_formlogin"),
                *pairs,
                ("groupingTable:xls", "groupingTable:xls"),
                ("javax.faces.ViewState", self.view_state),
            ],
            timeout=120,
        )
        return parse_vahan_workbook(workbook_content, segment_id, year)

    def _post_partial(
        self,
        source: str,
        execute: str,
        render: str,
        pairs: list[tuple[str, str]],
        timeout: int = 30,
    ) -> str:
        payload = [
            ("masterLayout_formlogin", "masterLayout_formlogin"),
            ("javax.faces.partial.ajax", "true"),
            ("javax.faces.source", source),
            ("javax.faces.partial.execute", execute),
            ("javax.faces.partial.render", render),
            *pairs,
            ("javax.faces.ViewState", self.view_state),
        ]
        response_text = self._post_text(
            self.action_url,
            headers=self.ajax_headers,
            pairs=payload,
            timeout=timeout,
        )
        self.view_state = self._extract_view_state_from_partial(response_text)
        return response_text

    def _get_text(self, url: str, headers: dict[str, str], timeout: int) -> str:
        if self.transport == "curl":
            return self._curl_text(url, headers=headers, timeout=timeout)

        response = self.session.get(url, headers=headers, timeout=timeout)
        response.raise_for_status()
        return response.text

    def _post_text(
        self,
        url: str,
        headers: dict[str, str],
        pairs: list[tuple[str, str]],
        timeout: int,
    ) -> str:
        if self.transport == "curl":
            return self._curl_text(url, headers=headers, data=pairs, timeout=timeout)

        response = self.session.post(url, headers=headers, data=pairs, timeout=timeout)
        response.raise_for_status()
        return response.text

    def _post_bytes(
        self,
        url: str,
        headers: dict[str, str],
        pairs: list[tuple[str, str]],
        timeout: int,
    ) -> bytes:
        if self.transport == "curl":
            return self._curl_bytes(url, headers=headers, data=pairs, timeout=timeout)

        response = self.session.post(url, headers=headers, data=pairs, timeout=timeout)
        response.raise_for_status()
        return response.content

    def _curl_text(
        self,
        url: str,
        headers: dict[str, str],
        data: list[tuple[str, str]] | None = None,
        timeout: int = 30,
    ) -> str:
        result = self._run_curl(url, headers=headers, data=data, timeout=timeout)
        return result.decode("utf-8", errors="ignore")

    def _curl_bytes(
        self,
        url: str,
        headers: dict[str, str],
        data: list[tuple[str, str]] | None = None,
        timeout: int = 30,
    ) -> bytes:
        return self._run_curl(url, headers=headers, data=data, timeout=timeout)

    def _run_curl(
        self,
        url: str,
        headers: dict[str, str],
        data: list[tuple[str, str]] | None,
        timeout: int,
    ) -> bytes:
        command = [
            "curl.exe" if os.name == "nt" else "curl",
            "--silent",
            "--show-error",
            "--location",
            "--compressed",
            "--max-time",
            str(max(timeout, 90)),  # Parivahan from cloud IPs is slow; give it room
            "--connect-timeout",
            "30",
            "--retry",
            "2",
            "--retry-delay",
            "5",
            "--retry-connrefused",
        ]
        if self.cookie_jar_path:
            command.extend(["-b", str(self.cookie_jar_path), "-c", str(self.cookie_jar_path)])
        for key, value in headers.items():
            command.extend(["-H", f"{key}: {value}"])
        if data:
            command.extend(["--data", urlencode(data, doseq=True)])
        command.append(url)
        completed = subprocess.run(command, capture_output=True, check=False)
        if completed.returncode != 0:
            stderr = completed.stderr.decode("utf-8", errors="ignore").strip()
            raise RuntimeError(stderr or f"curl failed for {url}")
        return completed.stdout

    @staticmethod
    def _extract_view_state_from_html(html: str) -> str:
        match = re.search(r'name="javax.faces.ViewState"[^>]*value="([^"]+)"', html)
        if not match:
            raise ValueError("Parivahan page did not expose a JSF view state.")
        return match.group(1)

    @staticmethod
    def _extract_view_state_from_partial(xml_text: str) -> str:
        match = re.search(
            r'<update id="j_id1:javax.faces.ViewState:0"><!\[CDATA\[(.*?)\]\]></update>',
            xml_text,
            flags=re.DOTALL,
        )
        if not match:
            raise ValueError("Parivahan partial response did not include a JSF view state.")
        return match.group(1)

    @staticmethod
    def _extract_form_action(html: str) -> str:
        match = re.search(r'<form[^>]+action="([^"]+)"', html)
        if not match:
            raise ValueError("Parivahan page did not expose the report form action.")
        return urljoin("https://vahan.parivahan.gov.in", match.group(1))


def parse_vahan_workbook(content: bytes, segment_id: str, year: int) -> dict[str, Any]:
    import tempfile
    from zipfile import BadZipFile

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as handle:
        handle.write(content)
        temp_path = Path(handle.name)

    try:
        workbook = load_workbook(temp_path, data_only=True, read_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    except BadZipFile as exc:
        debug_path = Path("data") / f"debug_vahan_invalid_{segment_id}_{year}.html"
        debug_path.write_bytes(content)
        raise ValueError(f"Parivahan workbook response for {segment_id} {year} was not an xlsx file. Saved response to {debug_path}.") from exc
    finally:
        temp_path.unlink(missing_ok=True)

    if len(rows) < 5:
        raise ValueError(f"Parivahan workbook for {segment_id} {year} was unexpectedly short.")

    month_headers = []
    for cell in rows[3][2:-1]:
        label = str(cell or "").strip().upper()
        if label in MONTH_NAME_TO_NUMBER:
            month_headers.append(label)

    maker_rows: list[dict[str, Any]] = []
    for row in rows[4:]:
        maker = str(row[1] or "").strip()
        if not maker:
            continue
        months: dict[str, int] = {}
        for index, month_name in enumerate(month_headers, start=2):
            raw_value = row[index] if index < len(row) else 0
            units = parse_number(raw_value)
            month_id = f"{year}-{MONTH_NAME_TO_NUMBER[month_name]:02d}"
            months[month_id] = units
        total_value = parse_number(row[2 + len(month_headers)] if 2 + len(month_headers) < len(row) else 0)
        maker_rows.append(
            {
                "maker": maker,
                "months": months,
                "total_units": total_value,
            }
        )

    return {
        "segment": segment_id,
        "year": year,
        "months": [f"{year}-{MONTH_NAME_TO_NUMBER[name]:02d}" for name in month_headers],
        "rows": maker_rows,
    }


def parse_number(value: Any) -> int:
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(round(value))
    cleaned = re.sub(r"[^\d.-]", "", str(value))
    if not cleaned:
        return 0
    return int(round(float(cleaned)))


def refresh_vahan_oem_cache(
    cache_path: Path,
    segment_ids: list[str] | None = None,
) -> dict[str, str | bool | None]:
    current_year = date.today().year
    years = list(range(current_year - 5, current_year + 1))
    cache_path.parent.mkdir(parents=True, exist_ok=True)

    existing: dict[str, Any] = {}
    if cache_path.exists():
        try:
            existing = json.loads(cache_path.read_text(encoding="utf-8"))
        except Exception:
            existing = {}

    requested_segments = segment_ids or list(SEGMENT_FILTERS.keys())
    target_years = years
    updated = False
    cache = existing if existing else {
        "generated_at": None,
        "source_url": VAHAN_ANALYTICS_URL,
        "validation_url": VAHAN_REPORT_URL,
        "segments": {},
    }

    try:
        client = VahanOemClient()
        for segment_id in requested_segments:
            segment_config = SEGMENT_FILTERS[segment_id]
            segment_bucket = cache["segments"].setdefault(
                segment_id,
                {
                    "label": segment_config["label"],
                    "group": segment_config["group"],
                    "years": {},
                },
            )
            existing_years = {int(year) for year in segment_bucket.get("years", {}).keys()}
            years_to_refresh = target_years if not existing_years else [year for year in target_years if year == current_year or year not in existing_years]
            for year in years_to_refresh:
                dataset = client.fetch_segment_year(segment_id, year)
                segment_bucket["years"][str(year)] = dataset
                updated = True

        cache["generated_at"] = datetime.now(UTC).isoformat()
        cache["source_url"] = VAHAN_ANALYTICS_URL
        cache["validation_url"] = VAHAN_REPORT_URL
        cache_path.write_text(json.dumps(cache, indent=2), encoding="utf-8")

        latest_month = f"{current_year}-{date.today().month:02d}"
        message = "Official Parivahan maker-wise month exports refreshed for OEM tracker tables."
        return VahanCacheResult(
            source="Vahan OEM tracker",
            status="ok",
            message=message,
            updated=updated,
            latest_value=latest_month,
        ).to_dict()
    except Exception as exc:
        return VahanCacheResult(
            source="Vahan OEM tracker",
            status="warning",
            message=f"Official Parivahan OEM refresh failed ({exc}). Retained the last validated Vahan OEM cache if available.",
            updated=False,
            latest_value=existing.get("generated_at"),
        ).to_dict()


def read_vahan_oem_cache(cache_path: Path) -> dict[str, Any] | None:
    if not cache_path.exists():
        return None
    try:
        return json.loads(cache_path.read_text(encoding="utf-8"))
    except Exception:
        return None
