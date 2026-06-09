"""Refresh data/raw_material_history.json with monthly commodity prices.

Two source paths, in order of preference:

1. **Local World Bank Pink Sheet Excel** at data/raw_material_pink_sheet.xlsx
   (CMO-Historical-Data-Monthly.xlsx). Single file, all 6 commodities,
   monthly history from 1960-present. Refresh by manually re-downloading
   the latest "Monthly prices (Pink Sheet) — Excel" from
   https://www.worldbank.org/en/research/commodity-markets and committing
   it. Authoritative — IMF/World Bank publish FRED's PAL/PCOPP/etc series
   from this same source.

2. **FRED CSV** (per-series fetch from https://fred.stlouisfed.org/). Used
   as a fallback when the Pink Sheet xlsx is missing. Note: FRED has
   started returning 403 to GitHub-Actions runners, so this path may not
   work reliably from CI — the Pink Sheet xlsx is the recommended primary.

Series produced (all monthly, USD-denominated):
  aluminum       — USD / metric ton
  copper         — USD / metric ton
  nickel         — USD / metric ton
  lead           — USD / metric ton
  iron_ore       — USD / dry metric ton (62% Fe CFR China benchmark)
  natural_rubber — USD / kg (TSR20 from Pink Sheet, RSS3-converted from FRED)

Output schema (data/raw_material_history.json):
  {
    "schema_version": 1,
    "as_of_date": "YYYY-MM-DD",
    "source_name": "...",
    "source_url": "...",
    "materials": [
      {
        "id": "aluminum", "label": "Aluminum",
        "unit_label": "USD / metric ton", "axis_suffix": "/mt",
        "wb_code"|"fred_series": "...",
        "series": [{"period": "YYYY-MM", "label": "Mon YYYY", "value": ...}, ...]
      },
      ...
    ]
  }

Window: clipped to START_PERIOD onwards (default 2016-01) to keep the file
slim. Bump back if a future analysis wants longer history.
"""
from __future__ import annotations

import csv
import io
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import requests


HISTORY_PATH = Path("data/raw_material_history.json")
PINK_SHEET_PATH = Path("data/raw_material_pink_sheet.xlsx")
# World Bank refreshes this canonical monthly workbook in place each month, so
# downloading it yields the latest data without the old manual "re-download and
# commit the xlsx" step (which is why raw materials silently froze on whatever
# month the committed copy happened to hold).
PINK_SHEET_URL = (
    "https://thedocs.worldbank.org/en/doc/"
    "5d903e848db1d1b83e0ec8f744e55570-0350012021/related/CMO-Historical-Data-Monthly.xlsx"
)
TIMEOUT = 30
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/136.0 Safari/537.36"
    ),
    "Accept": "text/csv,*/*;q=0.8",
}

# Earliest month to keep. World Bank Pink Sheet has data back to 1960; FRED to
# the late 1980s. 2016-01 gives us 10+ years which covers the dashboard's
# "show me the COVID dip and recovery" use case without bloating the JSON.
START_PERIOD = "2016-01"  # YYYY-MM in our internal format

FRED_CSV_URL = "https://fred.stlouisfed.org/graph/fredgraph.csv?id={series}"

# Pink Sheet column indices in the "Monthly Prices" sheet (verified May 2026
# release). If World Bank ever reorders, re-derive by scanning row 4 for the
# commodity name.
PINK_SHEET_COLS = {
    "aluminum": 62,
    "copper": 64,
    "nickel": 67,
    "lead": 65,
    "iron_ore": 63,
    "natural_rubber": 55,  # Rubber, TSR20 — already USD/kg
}

# Common metadata for both source paths. Conversion only used by FRED rubber
# (PRUBBUSDM is RSS3 cents/lb; Pink Sheet TSR20 is already USD/kg).
MATERIALS = [
    {"id": "aluminum",       "label": "Aluminum",                 "unit_label": "USD / metric ton",     "axis_suffix": "/mt",  "fred_series": "PALUMUSDM",   "wb_code": "ALUMINUM",     "convert": None},
    {"id": "copper",         "label": "Copper",                   "unit_label": "USD / metric ton",     "axis_suffix": "/mt",  "fred_series": "PCOPPUSDM",   "wb_code": "COPPER",       "convert": None},
    {"id": "nickel",         "label": "Nickel",                   "unit_label": "USD / metric ton",     "axis_suffix": "/mt",  "fred_series": "PNICKUSDM",   "wb_code": "NICKEL",       "convert": None},
    {"id": "lead",           "label": "Lead",                     "unit_label": "USD / metric ton",     "axis_suffix": "/mt",  "fred_series": "PLEADUSDM",   "wb_code": "LEAD",         "convert": None},
    {"id": "iron_ore",       "label": "Iron Ore (steel proxy)",   "unit_label": "USD / dry metric ton", "axis_suffix": "/dmt", "fred_series": "PIORECRUSDM", "wb_code": "IRON_ORE",     "convert": None},
    {"id": "natural_rubber", "label": "Natural Rubber (TSR20)",   "unit_label": "USD / kg",             "axis_suffix": "/kg",  "fred_series": "PRUBBUSDM",   "wb_code": "RUBBER_TSR20", "convert_fred_only": "rubber_cents_lb_to_usd_kg"},
]


# ─────────────────────────── World Bank Pink Sheet ────────────────────────────


def _pink_sheet_period_to_ymd(period_str) -> str | None:
    # Pink Sheet uses "YYYYMmm" e.g. "2016M01"
    m = re.match(r"^(\d{4})M(\d{2})$", str(period_str))
    return f"{m.group(1)}-{m.group(2)}" if m else None


def _download_pink_sheet() -> bytes | None:
    """Fetch the latest Pink Sheet workbook from World Bank. Returns the xlsx
    bytes, or None if the download isn't usable (caller falls back to the
    committed local copy)."""
    try:
        resp = requests.get(PINK_SHEET_URL, headers=HEADERS, timeout=TIMEOUT, allow_redirects=True)
    except Exception as exc:
        print(f"  Pink Sheet download failed: {exc}", flush=True)
        return None
    if resp.status_code != 200 or resp.content[:2] != b"PK":
        print(
            f"  Pink Sheet download not usable (HTTP {resp.status_code}, "
            f"{len(resp.content)} bytes); using local copy.",
            flush=True,
        )
        return None
    print(f"  Downloaded latest Pink Sheet ({len(resp.content):,} bytes) from World Bank.", flush=True)
    return resp.content


def _read_pink_sheet() -> dict[str, list[dict]] | None:
    """Returns {material_id: [{period, label, value}, ...]} from the World Bank
    Pink Sheet. Prefers a freshly downloaded workbook so the series self-updates
    each month; falls back to the committed PINK_SHEET_PATH when the download is
    unavailable. Returns None if neither source / openpyxl is available, or the
    workbook structure isn't what we expect."""
    try:
        from openpyxl import load_workbook  # noqa: WPS433
    except ImportError:
        print("  openpyxl not installed; can't read Pink Sheet Excel.", flush=True)
        return None

    blob = _download_pink_sheet()
    source = "World Bank download"
    if blob is None:
        if not PINK_SHEET_PATH.exists():
            return None
        blob = PINK_SHEET_PATH.read_bytes()
        source = "local committed copy"

    try:
        wb = load_workbook(io.BytesIO(blob), data_only=True, read_only=True)
    except Exception as exc:
        print(f"  Pink Sheet xlsx read failed ({source}): {exc}", flush=True)
        # A corrupt download shouldn't lose us the local fallback.
        if source != "local committed copy" and PINK_SHEET_PATH.exists():
            try:
                wb = load_workbook(io.BytesIO(PINK_SHEET_PATH.read_bytes()), data_only=True, read_only=True)
                source = "local committed copy"
            except Exception as exc2:
                print(f"  Local Pink Sheet read also failed: {exc2}", flush=True)
                return None
        else:
            return None
    print(f"  Pink Sheet source: {source}.", flush=True)
    if "Monthly Prices" not in wb.sheetnames:
        print(f"  Pink Sheet missing 'Monthly Prices' tab (got: {wb.sheetnames})", flush=True)
        return None
    ws = wb["Monthly Prices"]

    buckets: dict[str, list[dict]] = {mid: [] for mid in PINK_SHEET_COLS}
    started = False
    for row in ws.iter_rows(min_row=8, values_only=True):
        period_raw = row[0]
        if not period_raw:
            continue
        ym = _pink_sheet_period_to_ymd(period_raw)
        if not ym:
            continue
        if ym >= START_PERIOD:
            started = True
        if not started:
            continue
        for mid, col in PINK_SHEET_COLS.items():
            v = row[col] if col < len(row) else None
            if isinstance(v, (int, float)):
                buckets[mid].append({
                    "period": ym,
                    "label": _month_label(ym),
                    "value": round(float(v), 4),
                })
    return buckets if any(buckets.values()) else None


# ─────────────────────────────── FRED fallback ────────────────────────────────


def _fred_url(series: str) -> str:
    return FRED_CSV_URL.format(series=series)


def _fetch_fred_series(series_id: str) -> list[tuple[str, float]]:
    """Returns chronological [(YYYY-MM, value), ...] pairs from FRED."""
    url = _fred_url(series_id)
    try:
        response = requests.get(url, headers=HEADERS, timeout=TIMEOUT, allow_redirects=True)
        response.raise_for_status()
    except Exception as exc:
        print(f"  FRED fetch failed for {series_id}: {exc}", flush=True)
        return []
    reader = csv.reader(io.StringIO(response.text))
    rows = list(reader)
    if not rows or len(rows) < 2:
        print(f"  FRED returned empty CSV for {series_id}", flush=True)
        return []
    out: list[tuple[str, float]] = []
    for row in rows[1:]:
        if len(row) < 2:
            continue
        date_str, val_str = row[0], row[1]
        if val_str == "." or not val_str:
            continue
        try:
            month_id = date_str[:7]  # YYYY-MM
            value = float(val_str)
        except (ValueError, IndexError):
            continue
        out.append((month_id, value))
    return out


def _convert_value(value: float, mode: str | None) -> float:
    if mode is None:
        return value
    if mode == "rubber_cents_lb_to_usd_kg":
        return value * 0.01 * 2.20462
    return value


# ─────────────────────────────── Common helpers ───────────────────────────────


def _month_label(month_id: str) -> str:
    try:
        d = datetime.strptime(month_id, "%Y-%m")
    except ValueError:
        return month_id
    return d.strftime("%b %Y")


def _build_payload(materials_out: list[dict], source_name: str, source_url: str, source_note: str) -> dict:
    return {
        "schema_version": 1,
        "as_of_date": datetime.now(tz=timezone.utc).strftime("%Y-%m-%d"),
        "source_name": source_name,
        "source_url": source_url,
        "source_note": source_note,
        "materials": materials_out,
    }


# ──────────────────────────────────── Main ────────────────────────────────────


def _try_pink_sheet() -> dict | None:
    print("--- Source 1: World Bank Pink Sheet (data/raw_material_pink_sheet.xlsx) ---", flush=True)
    buckets = _read_pink_sheet()
    if not buckets:
        print("  Pink Sheet not available; falling through to FRED.", flush=True)
        return None
    materials_out = []
    for spec in MATERIALS:
        series = buckets.get(spec["id"], [])
        if not series:
            continue
        print(
            f"  {spec['id']}: {len(series)} months "
            f"({series[0]['period']} → {series[-1]['period']}); "
            f"latest {series[-1]['value']:,.2f} {spec['axis_suffix']}",
            flush=True,
        )
        materials_out.append({
            "id": spec["id"],
            "label": spec["label"],
            "unit_label": spec["unit_label"],
            "axis_suffix": spec["axis_suffix"],
            "wb_code": spec["wb_code"],
            "series": series,
        })
    if not materials_out:
        return None
    return _build_payload(
        materials_out,
        "World Bank Commodity Markets — Pink Sheet (Monthly Prices)",
        "https://www.worldbank.org/en/research/commodity-markets",
        (
            "Monthly nominal USD commodity prices ingested from the local "
            "data/raw_material_pink_sheet.xlsx (World Bank CMO-Historical-Data-Monthly). "
            "Aluminum / Copper / Nickel / Lead are LME spot. Iron Ore is the "
            "62% Fe CFR China benchmark. Natural Rubber is TSR20 USD/kg."
        ),
    )


def _try_fred() -> dict | None:
    print("--- Source 2: FRED CSV (https://fred.stlouisfed.org/) ---", flush=True)
    materials_out = []
    failed: list[str] = []
    for spec in MATERIALS:
        print(f"  fetching {spec['id']} ({spec['fred_series']})...", flush=True)
        raw_pairs = _fetch_fred_series(spec["fred_series"])
        if not raw_pairs:
            failed.append(spec["id"])
            continue
        clipped = [(m, v) for (m, v) in raw_pairs if m >= START_PERIOD]
        convert = spec.get("convert_fred_only") or spec.get("convert")
        series = [
            {"period": m, "label": _month_label(m), "value": round(_convert_value(v, convert), 4)}
            for (m, v) in clipped
        ]
        if not series:
            failed.append(spec["id"])
            continue
        print(
            f"  {spec['id']}: {len(series)} months "
            f"({series[0]['period']} → {series[-1]['period']}); "
            f"latest {series[-1]['value']:,.2f} {spec['axis_suffix']}",
            flush=True,
        )
        materials_out.append({
            "id": spec["id"],
            "label": spec["label"],
            "unit_label": spec["unit_label"],
            "axis_suffix": spec["axis_suffix"],
            "fred_series": spec["fred_series"],
            "series": series,
        })
    if not materials_out:
        return None
    return _build_payload(
        materials_out,
        "FRED (St. Louis Fed) — IMF / World Bank Commodity Prices",
        "https://fred.stlouisfed.org/",
        (
            "Monthly commodity prices fetched from the FRED public CSV endpoint. "
            "Aluminum / Copper / Nickel / Lead / Iron Ore are USD-per-metric-ton "
            "series. Natural Rubber is FRED's RSS3 cents-per-pound (PRUBBUSDM), "
            "converted to USD/kg."
        ),
    )


def main() -> int:
    payload = _try_pink_sheet() or _try_fred()
    if not payload:
        print("\nBoth sources failed; not writing history file.", file=sys.stderr)
        return 1

    HISTORY_PATH.parent.mkdir(parents=True, exist_ok=True)
    HISTORY_PATH.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    print(
        f"\nWrote {HISTORY_PATH} ({len(payload['materials'])} materials, "
        f"source: {payload['source_name']}).",
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
