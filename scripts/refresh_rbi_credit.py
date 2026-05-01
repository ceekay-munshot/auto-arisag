"""Refresh RBI Vehicle Loans + Non-food Bank Credit data into
``data/rbi_credit.json`` by walking RBI's official "Data on Sectoral
Deployment of Bank Credit" listing.

Path the script follows (mirrors what a human does):

    1. GET https://rbi.org.in/Scripts/Data_Sectoral_Deployment.aspx
       This page lists every monthly release as a row with a date heading
       (e.g. "Apr 30, 2026") followed by a press-release link
       ("Sectoral Deployment of Bank Credit – March 2026" pointing to
       BS_PressReleaseDisplay.aspx?prid=NNNNN).

    2. For each release whose reference month is not already in
       ``data/rbi_credit.json``, GET the detail page and find the link
       to "Statements I and II" — that's an .xlsx file
       (typically ``/upload/PressReleases/Excel/SIBCS<DDMMYYYY>.xlsx``).

    3. Download the Excel and read Statement 1 ("Deployment of Gross
       Bank Credit by Major Sectors"). Two rows matter to us:
         - ``III. Non-food Credit`` (top of the statement)
         - ``4.7 Vehicle Loans`` (under Personal Loans)
       The rightmost "Outstanding as on" column is the current month
       observation; the rightmost YoY% column is the printed YoY growth.

    4. Append a row ``{month, as_of_date, outstanding_cr, yoy_pct,
       non_food_total_cr, non_food_yoy_pct, source_url}`` and save.
       Idempotent.

Run via the GitHub Actions cron — sandbox networks cannot reach RBI.
"""

from __future__ import annotations

import calendar
import html
import json
import re
import sys
from datetime import datetime
from io import BytesIO
from pathlib import Path
from urllib.parse import urljoin

import requests
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from dashboard.update_snapshot import (  # noqa: E402
    MONTH_LOOKUP,
    to_month_id,
)


HISTORY_PATH = Path("data/rbi_credit.json")
LISTING_URL = "https://rbi.org.in/Scripts/Data_Sectoral_Deployment.aspx"
TIMEOUT = 60
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/136.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-IN,en;q=0.9",
}
EXCEL_HEADERS = {
    **HEADERS,
    "Accept": (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,"
        "application/vnd.ms-excel,*/*;q=0.8"
    ),
}

# Step 1: pick every <a> whose href points at a press release (prid=N).
# We capture (href, inner_html) and decode the inner text afterwards. This
# is far more robust than trying to anchor the title inside the same regex
# because RBI sometimes wraps the title in nested font/span tags or uses
# HTML entities (`&ndash;`) for the dash.
PRESS_RELEASE_ANCHOR_RE = re.compile(
    r'<a[^>]+href="([^"]*BS_PressReleaseDisplay\.aspx\?[^"]*?prid=\d+[^"]*)"[^>]*>'
    r"(.*?)</a>",
    re.IGNORECASE | re.DOTALL,
)

# Step 2: out of the anchor texts, pull the (month, year) from the title.
SDBC_TITLE_RE = re.compile(
    r"Sectoral\s+Deployment\s+of\s+Bank\s+Credit\b"
    r"[^A-Za-z]{1,8}"  # tolerate any combination of dashes / spaces / punctuation
    r"(January|February|March|April|May|June|July|August|September|October|November|December)"
    r"[\s,]*(\d{4})",
    re.IGNORECASE,
)

# Detail page → first .xlsx attachment. RBI labels the link "Statements
# I and II" but we just match the file extension to be tolerant of label
# changes.
EXCEL_HREF_RE = re.compile(
    r'<a[^>]+href="([^"]+\.xlsx)"',
    re.IGNORECASE,
)


def _fetch_html(url: str) -> str | None:
    try:
        response = requests.get(url, headers=HEADERS, timeout=TIMEOUT, allow_redirects=True)
        response.raise_for_status()
        return response.text
    except Exception as exc:
        print(f"  fetch html failed {url}: {exc}", flush=True)
        return None


def _fetch_excel(url: str) -> bytes | None:
    try:
        response = requests.get(url, headers=EXCEL_HEADERS, timeout=TIMEOUT, allow_redirects=True)
        response.raise_for_status()
        if not response.content:
            return None
        # OOXML magic bytes (zip archive). Helps us reject HTML error pages
        # served with a 200 status code.
        if not response.content.startswith(b"PK"):
            print(f"  {url}: not an xlsx (first bytes: {response.content[:8]!r})", flush=True)
            return None
        return response.content
    except Exception as exc:
        print(f"  fetch excel failed {url}: {exc}", flush=True)
        return None


def _anchor_text(inner_html: str) -> str:
    """Strip nested tags and decode HTML entities so we can match the
    title with a plain regex."""
    text = re.sub(r"<[^>]+>", " ", inner_html)
    text = html.unescape(text)
    return re.sub(r"\s+", " ", text).strip()


def _discover_releases(listing_html: str) -> list[tuple[str, str]]:
    """Return ``[(month_id, detail_url), ...]`` for every release the
    listing page advertises. ``month_id`` is the *reference* month (e.g.
    Mar 2026 for the release titled "March 2026" that drops in Apr 2026)."""
    out: list[tuple[str, str]] = []
    seen: set[str] = set()
    anchors = PRESS_RELEASE_ANCHOR_RE.findall(listing_html)
    print(f"  press-release anchors found: {len(anchors)}", flush=True)
    for relative_href, inner in anchors:
        text = _anchor_text(inner)
        match = SDBC_TITLE_RE.search(text)
        if not match:
            continue
        month_name, year_text = match.groups()
        try:
            month_id = to_month_id(month_name, int(year_text))
        except ValueError:
            continue
        detail_url = urljoin(LISTING_URL, relative_href)
        if detail_url in seen:
            continue
        seen.add(detail_url)
        out.append((month_id, detail_url))
    if not out and anchors:
        # Surface the first few anchor texts so the workflow log tells us
        # exactly why our title regex didn't match (encoding shift, layout
        # change, etc.) without needing another round-trip.
        sample = [_anchor_text(inner)[:120] for _, inner in anchors[:8]]
        print(f"  no SDBC titles matched. First anchor texts: {sample}", flush=True)
    return out


def _excel_attachment_url(detail_html: str, detail_url: str) -> str | None:
    match = EXCEL_HREF_RE.search(detail_html)
    if not match:
        return None
    return urljoin(detail_url, match.group(1))


_VEHICLE_LABEL_RE = re.compile(r"(?:^|\s)(?:4\.7\.?\s*)?Vehicle\s+Loans?\b", re.IGNORECASE)
_NON_FOOD_LABEL_RE = re.compile(r"(?:III\.?\s*)?Non[-\s]?food\s+Credit\b", re.IGNORECASE)


def _parse_excel(content: bytes) -> dict[str, float | int | None] | None:
    """Pull (vehicle_outstanding, vehicle_yoy, non_food_outstanding,
    non_food_yoy) from the Excel workbook's "Statement 1" sheet.

    Implementation notes:
      - We avoid hard-coded row indices because RBI tweaks layout slightly
        between years. Instead we scan column A for the labels we want and
        then read rightmost-numeric "Outstanding as on" + rightmost YoY%
        cells in the same row.
      - "Outstanding as on" cells are large integers (Rs. crore).
      - "Variation (Year-on-Year) %" cells are floats with one decimal,
        usually < 100.
    """
    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    statement_sheet = None
    for name in workbook.sheetnames:
        if "statement" in name.lower() and ("1" in name or "i" in name.lower()):
            statement_sheet = workbook[name]
            break
    if statement_sheet is None:
        statement_sheet = workbook[workbook.sheetnames[0]]

    veh_row: list = []
    nf_row: list = []
    for row in statement_sheet.iter_rows(values_only=True):
        if not row:
            continue
        first_cell = row[0]
        if not isinstance(first_cell, str):
            continue
        label = first_cell.strip()
        if not nf_row and _NON_FOOD_LABEL_RE.search(label):
            nf_row = list(row)
        elif not veh_row and _VEHICLE_LABEL_RE.search(label):
            veh_row = list(row)
        if veh_row and nf_row:
            break

    if not veh_row or not nf_row:
        return None

    veh_outstanding, veh_yoy = _last_outstanding_and_yoy(veh_row)
    nf_outstanding, nf_yoy = _last_outstanding_and_yoy(nf_row)
    return {
        "vehicle_outstanding_cr": veh_outstanding,
        "vehicle_yoy_pct": veh_yoy,
        "non_food_total_cr": nf_outstanding,
        "non_food_yoy_pct": nf_yoy,
    }


def _last_outstanding_and_yoy(row: list) -> tuple[int | None, float | None]:
    """In an RBI Statement-1 row, every numeric > ~50,000 is an "Outstanding"
    cell (Rs crore), and every numeric in [-100, 200] (rounded one decimal)
    is a YoY/FY variation. The right-most outstanding is the latest
    observation; the right-most variation is the YoY% RBI printed."""
    outstanding = None
    yoy = None
    for cell in row[1:]:
        if cell is None:
            continue
        try:
            value = float(cell)
        except (TypeError, ValueError):
            continue
        if abs(value) > 50_000 and float(int(value)) == value:
            outstanding = int(value)
        elif -100 <= value <= 200:
            yoy = round(value, 2)
    return outstanding, yoy


def _last_reporting_friday(month_id: str) -> str:
    """Return ISO date of the last Friday of ``month_id``. Used as a
    reasonable ``as_of_date`` fallback when we can't read the actual
    reporting Friday from the spreadsheet header."""
    year, month = (int(part) for part in month_id.split("-"))
    last_day = calendar.monthrange(year, month)[1]
    cur = datetime(year, month, last_day)
    while cur.weekday() != 4:  # 4 = Friday
        cur = cur.replace(day=cur.day - 1)
    return cur.strftime("%Y-%m-%d")


def _load_history() -> dict:
    if HISTORY_PATH.exists():
        try:
            return json.loads(HISTORY_PATH.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            pass
    return {
        "source_name": "RBI — Sectoral Deployment of Bank Credit",
        "source_url": LISTING_URL,
        "concept": "Outstanding scheduled commercial bank credit for Vehicle Loans (sub-segment of Personal Loans).",
        "unit": "INR crore (outstanding)",
        "coverage_note": "",
        "is_seed": True,
        "series": [],
    }


def _save_history(payload: dict) -> None:
    HISTORY_PATH.parent.mkdir(parents=True, exist_ok=True)
    HISTORY_PATH.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def main() -> int:
    listing_html = _fetch_html(LISTING_URL)
    if not listing_html:
        print("RBI listing fetch failed; nothing to update.", file=sys.stderr)
        return 1
    print(f"RBI listing: HTML length {len(listing_html)} chars", flush=True)
    releases = _discover_releases(listing_html)
    if not releases:
        # Last-ditch breadcrumb: dump the first ~600 chars of body so we
        # can eyeball whether RBI returned a real listing or some
        # interstitial / error page.
        snippet = re.sub(r"\s+", " ", listing_html)[:600]
        print(f"  listing snippet: {snippet}", flush=True)
        print("RBI listing returned 0 releases.", file=sys.stderr)
        return 1
    print(f"RBI listing: discovered {len(releases)} releases", flush=True)

    history = _load_history()
    by_month: dict[str, dict] = {row["month"]: row for row in history.get("series", []) if row.get("month")}

    added = 0
    refreshed = 0
    skipped = 0
    parse_failed = 0

    for month_id, detail_url in releases:
        existing = by_month.get(month_id)
        if existing and existing.get("source_url", "").startswith("https://www.rbi.org.in"):
            skipped += 1
            continue
        detail_html = _fetch_html(detail_url)
        if not detail_html:
            continue
        excel_url = _excel_attachment_url(detail_html, detail_url)
        if not excel_url:
            print(f"  {month_id}: no .xlsx attachment found at {detail_url}", flush=True)
            parse_failed += 1
            continue
        excel_bytes = _fetch_excel(excel_url)
        if not excel_bytes:
            parse_failed += 1
            continue
        parsed = _parse_excel(excel_bytes)
        if not parsed or not parsed.get("vehicle_outstanding_cr"):
            print(f"  {month_id}: workbook parsed but Vehicle Loans row missing", flush=True)
            parse_failed += 1
            continue

        record = {
            "month": month_id,
            "as_of_date": _last_reporting_friday(month_id),
            "outstanding_cr": parsed["vehicle_outstanding_cr"],
            "yoy_pct": parsed["vehicle_yoy_pct"],
            "non_food_total_cr": parsed["non_food_total_cr"],
            "non_food_yoy_pct": parsed["non_food_yoy_pct"],
            "source_url": detail_url,
        }
        if existing and existing == record:
            continue
        if existing:
            refreshed += 1
        else:
            added += 1
        by_month[month_id] = record
        print(
            f"  {month_id}: VL ₹{record['outstanding_cr']:,} cr ({record['yoy_pct']}% YoY), "
            f"NF ₹{record['non_food_total_cr']:,} cr ({record['non_food_yoy_pct']}% YoY)",
            flush=True,
        )

    history["series"] = [by_month[k] for k in sorted(by_month)]
    history["source_url"] = LISTING_URL
    if added or refreshed:
        history["is_seed"] = False
        history["coverage_note"] = (
            "Outstanding amount as of the last reporting Friday of each month, "
            "parsed from RBI's Sectoral Deployment of Bank Credit Excel statements."
        )
    _save_history(history)

    print(
        f"\nWrote {HISTORY_PATH} (added {added}, refreshed {refreshed}, "
        f"skipped {skipped}, parse_failed {parse_failed}). "
        f"Total months: {len(history['series'])}",
        flush=True,
    )
    return 0 if (added + refreshed) > 0 or parse_failed == 0 else 2


if __name__ == "__main__":
    sys.exit(main())
